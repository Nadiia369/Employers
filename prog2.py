import csv
from datetime import datetime
import openpyxl

# Функція для розрахунку віку
def calculate_age(birth_date):
    today = datetime.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

# Створення XLSX файлу
def create_xlsx(csv_filename, xlsx_filename):
    try:
        # Створюємо Excel файл з 5 аркушами
        wb = openpyxl.Workbook()
        ws_all = wb.active
        ws_all.title = "all"
        ws_younger_18 = wb.create_sheet(title="younger_18")
        ws_18_45 = wb.create_sheet(title="18-45")
        ws_45_70 = wb.create_sheet(title="45-70")
        ws_older_70 = wb.create_sheet(title="older_70")

        headers = ['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік']
        for ws in [ws_all, ws_younger_18, ws_18_45, ws_45_70, ws_older_70]:
            ws.append(headers)

        # Читання даних з CSV
        with open(csv_filename, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Пропускаємо заголовки

            for idx, row in enumerate(reader, start=1):
                birth_date = datetime.strptime(row[4], "%Y-%m-%d")
                age = calculate_age(birth_date)

                # Записуємо всі дані в аркуш "all"
                ws_all.append([idx] + row[:3] + [row[4], age])

                # Записуємо дані у відповідний аркуш за віковими категоріями
                if age < 18:
                    ws_younger_18.append([idx] + row[:3] + [row[4], age])
                elif 18 <= age < 45:
                    ws_18_45.append([idx] + row[:3] + [row[4], age])
                elif 45 <= age < 70:
                    ws_45_70.append([idx] + row[:3] + [row[4], age])
                else:
                    ws_older_70.append([idx] + row[:3] + [row[4], age])

        # Збереження файлу
        wb.save(xlsx_filename)
        print("Ok, файл створено успішно.")
    except FileNotFoundError:
        print("Помилка: файл CSV не знайдено.")
    except Exception as e:
        print(f"Помилка при створенні XLSX: {e}")

# Створення XLSX файлу на основі CSV
create_xlsx('employees.csv', 'employees.xlsx')
